"""
calculate_effect_sizes.py
=========================
Reproducible script for computing Hedges' D and Odds Ratio effect sizes
for the Mate-Choice Copying (MCC) meta-analysis.

INPUT:
    Data Extraction Mate Choice Meta Analysis.xlsx
    Tab: EduardoExtraction

OUTPUT:
    Same file, columns updated in-place:
        AN (col 40): effectSizeCalculatedHedgesD
        AO (col 41): effectSizeCalculatedHedgesDVariance
        AP (col 42): effectSizeHedgesDHowCalculated
        AQ (col 43): effectSizeCalculatedOddsRatio
        AR (col 44): effectSizeCalculatedOddsRatioVariance
        AS (col 45): effectSizeOddsRatioHowCalculated

FORMULAS USED
-------------
Hedges' D (g) — bias-corrected standardised mean difference:

  Case 1: From group means and SDs
    SP     = sqrt[((n1-1)*SD1^2 + (n2-1)*SD2^2) / (n1+n2-2)]
    d      = (mean_treat - mean_ctrl) / SP
    J      = 1 - 3 / (4*(n1+n2-2) - 1)          # correction factor
    g      = d * J
    Var(g) = (n1+n2)/(n1*n2) + g^2 / (2*(n1+n2-2))

  Case 2: From a one-sample or paired t-statistic (df = n-1)
    d      = t / sqrt(n)
    J      = 1 - 3 / (4*(n-1) - 1)
    g      = d * J
    Var(g) = 1/n + g^2 / (2*(n-1))

  Case 3: From an independent-samples t-statistic (df = n1+n2-2)
    d      = t * sqrt(1/n1 + 1/n2)
    J      = 1 - 3 / (4*(n1+n2-2) - 1)
    g      = d * J
    Var(g) = (n1+n2)/(n1*n2) + g^2 / (2*(n1+n2-2))
    NOTE: When individual group sizes are unknown, n1 = n2 = N_total // 2.

  Missing SD recovered from SE:
    SD = SE * sqrt(n)

Odds Ratio (OR) — from a 2x2 contingency table:

  2x2 table layout:
    A = treatment group chose option associated with social information
    B = treatment group chose the alternative
    C = control group (or expected under H0) chose same option
    D = control group chose alternative

  OR       = (A * D) / (B * C)
  Var(logOR) = 1/A + 1/B + 1/C + 1/D       # delta method approximation

  50/50 rule: When no separate control group exists, C = D = n/2
  (i.e., the null expectation of 50% choosing each option).

DECISION RULES
--------------
  - Rows where effectSizeCalculatedHedgesD is already populated → SKIP
  - Rows where effectSizeCalculatedOddsRatio is already populated → SKIP
  - If A, B, C, D are all positive numbers → compute OR
  - If controlMean, controlSD/SE, controlN, treatmentMean, treatmentSD/SE,
    treatmentN are all available → compute Hedges' D from means
  - If estimateType is one-sample / paired t-test AND dataType is
    'Inferential statistics' → compute Hedges' D from t (Case 2)
  - If estimateType is independent t-test AND dataType is
    'Inferential statistics' → compute Hedges' D from t (Case 3)
  - F-statistics, chi-squared, GLMM outputs → SKIP (not directly convertible)

OR ↔ Hedges' D CONVERSION (Chinn 2000; Borenstein et al. 2009)
--------------------------------------------------------------
  These conversions are approximate (logistic–normal approximation).
  They are applied as a FINAL PASS after primary calculations:

  OR  → Hedges' g  (when OR is filled and Hedges' D is blank):
    d        = ln(OR) × (√3 / π)    ≈ ln(OR) × 0.5513
    Var(d)   = Var(logOR) × (3 / π²) ≈ Var(logOR) × 0.3040
    J        = 1 − 3 / (4 × df − 1)   where df = n − 1 or n1+n2−2
    g        = d × J
    Var(g)   ≈ Var(d) × J²             (approximate; see note)
    n source: N column (total) → fallback to A+B (treatment group)

  Hedges' g → OR  (when Hedges' D is filled and OR is blank):
    logOR    = g × (π / √3)          ≈ g × 1.8138
    OR       = exp(logOR)
    Var(logOR) = Var(g) × (π² / 3)   ≈ Var(g) × 3.2899

  howCalculated fields are annotated with "[converted from OR]" or
  "[converted from Hedges' D]" so converted values can be distinguished
  from directly-calculated ones.

KNOWN APPROXIMATIONS / NOTES
-----------------------------
  - Miller_2013: Welch t-test (df=40.2, N=48). Individual group sizes unknown;
    n1 = n2 = 24 assumed for g calculation.
  - Ophir_2009 Exp001-003: EstimateType labelled 't-statistic'; df = n-1
    confirms one-sample design. Treated as Case 2.
  - Kobayashi_2026: Before-after paired design but treated as independent
    groups (conservative). Values pre-populated from separate extraction.
  - Rows 16 (Huckvale), 48-49 (Nöbel ProcRoySocB), 62 (Ophir Exp004),
    64 (Plath Exp002): only F-statistics or chi-squared available. Skipped.
  - OR↔g conversions: J correction uses n from column N when available,
    else A+B (treatment count). If neither is available, J ≈ 1 (large n).

Author: Eduardo Santos (meta-analysis) / Claude (computation)
Date:   2026-05-26
"""

import math
import openpyxl

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

import os
EXCEL_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
    "data", "1_effect_size_calculation_pipeline", "Data Extraction Mate Choice Meta Analysis.xlsx"
)
SHEET_NAME = "EduardoExtraction"

# Mapping of logical keys to actual column header names in EduardoExtraction sheet.
# Resolving columns dynamically from headers allows inserting new columns (e.g., identifierEffectSizeID)
# without breaking the index mapping of existing variables.
HEADER_MAP = {
    "study":          "identifierStudyId",
    "exp":            "identifierExperimentId",
    "est_type":       "effectSizeEstimateType",
    "est_val":        "effectSizeEstimateValue",
    "est_extra":      "effectSizeEstimateExtra",
    "A":              "effectSizeSampleSizeA",
    "B":              "effectSizeSampleSizeB",
    "C":              "effectSizeSampleSizeC",
    "D":              "effectSizeSampleSizeD",
    "N":              "effectSizesampleSize",
    "ctrl_mean":      "controlMean",
    "ctrl_sd":        "controlSD",
    "ctrl_se":        "controlSE",
    "ctrl_n":         "controlN",
    "treat_mean":     "treatmentMean",
    "treat_sd":       "treatmentSD",
    "treat_se":       "treatmentSE",
    "treat_n":        "treatmentN",
    "data_type":      "effectSizeDataTypeForCalculation",
    "hedges_d":       "effectSizeCalculatedHedgesD",
    "hedges_d_var":   "effectSizeCalculatedHedgesDVariance",
    "hedges_d_how":   "effectSizeHedgesDHowCalculated",
    "or":             "effectSizeCalculatedOddsRatio",
    "or_var":         "effectSizeCalculatedOddsRatioVariance",
    "or_how":         "effectSizeOddsRatioHowCalculated",
}

# ---------------------------------------------------------------------------
# Core calculation functions
# ---------------------------------------------------------------------------

def hedges_g_from_means(m_treat, sd_treat, n_treat, m_ctrl, sd_ctrl, n_ctrl):
    """
    Compute Hedges' g from group means and standard deviations.

    Returns
    -------
    g : float   — bias-corrected effect size
    var_g : float — sampling variance of g
    """
    df = n_treat + n_ctrl - 2
    sp = math.sqrt(
        ((n_treat - 1) * sd_treat ** 2 + (n_ctrl - 1) * sd_ctrl ** 2) / df
    )
    d = (m_treat - m_ctrl) / sp
    J = 1 - 3 / (4 * df - 1)
    g = round(d * J, 4)
    var_g = round((n_treat + n_ctrl) / (n_treat * n_ctrl) + g ** 2 / (2 * df), 4)
    return g, var_g


def hedges_g_from_t_onesample(t, n):
    """
    Compute Hedges' g from a one-sample or paired t-statistic.
    Appropriate when df = n - 1 (same subjects measured twice, or
    single group compared to a fixed null).

    Returns
    -------
    g : float
    var_g : float
    """
    df = n - 1
    d = t / math.sqrt(n)
    J = 1 - 3 / (4 * df - 1)
    g = round(d * J, 4)
    var_g = round(1 / n + g ** 2 / (2 * df), 4)
    return g, var_g


def hedges_g_from_t_indep(t, n1, n2):
    """
    Compute Hedges' g from an independent-samples t-statistic.

    Returns
    -------
    g : float
    var_g : float
    """
    df = n1 + n2 - 2
    d = t * math.sqrt(1 / n1 + 1 / n2)
    J = 1 - 3 / (4 * df - 1)
    g = round(d * J, 4)
    var_g = round((n1 + n2) / (n1 * n2) + g ** 2 / (2 * df), 4)
    return g, var_g


def odds_ratio_2x2(A, B, C, D):
    """
    Compute the Odds Ratio and its variance from a 2x2 contingency table.

    Table layout:
        Chose option  Did not choose
    Treatment    A         B
    Control      C         D

    OR = (A * D) / (B * C)
    Var(log OR) = 1/A + 1/B + 1/C + 1/D

    Returns
    -------
    OR : float
    var_log_or : float   — variance of the LOG odds ratio
    """
    OR = round((A * D) / (B * C), 4)
    var_log_or = round(1 / A + 1 / B + 1 / C + 1 / D, 4)
    return OR, var_log_or


def or_to_hedges_g(OR, var_log_or, n=None):
    """
    Convert an Odds Ratio to Hedges' g using the logistic approximation
    (Chinn 2000; Borenstein et al. 2009).

    d        = ln(OR) × (√3 / π)
    Var(d)   = Var(logOR) × (3 / π²)
    J        = 1 − 3 / (4*df − 1)  where df = n − 1  (if n is known)
    g        = d × J
    Var(g)   ≈ Var(d) × J²

    Parameters
    ----------
    OR         : float — odds ratio (must be > 0)
    var_log_or : float — variance of ln(OR)
    n          : int or None — total sample size; if None, J ≈ 1 (large n)

    Returns
    -------
    g      : float — bias-corrected effect size (Hedges' g)
    var_g  : float — sampling variance of g
    j_used : float — J correction factor applied (1.0 if n is None)
    """
    log_or = math.log(OR)
    d      = log_or * (math.sqrt(3) / math.pi)
    var_d  = var_log_or * (3 / math.pi ** 2)

    if n is not None and n > 2:
        df = n - 1
        J  = 1 - 3 / (4 * df - 1)
    else:
        J  = 1.0   # large-sample approximation

    g     = round(d * J, 4)
    var_g = round(var_d * J ** 2, 4)
    return g, var_g, J


def hedges_g_to_or(g, var_g):
    """
    Convert Hedges' g to an Odds Ratio using the logistic approximation
    (Chinn 2000; Borenstein et al. 2009).

    logOR      = g × (π / √3)
    OR         = exp(logOR)
    Var(logOR) = Var(g) × (π² / 3)

    Parameters
    ----------
    g     : float — Hedges' g (bias-corrected standardised mean difference)
    var_g : float — sampling variance of g

    Returns
    -------
    OR         : float — odds ratio
    var_log_or : float — variance of ln(OR)
    """
    log_or     = g * (math.pi / math.sqrt(3))
    OR         = round(math.exp(log_or), 4)
    var_log_or = round(var_g * (math.pi ** 2 / 3), 4)
    return OR, var_log_or


# ---------------------------------------------------------------------------
# Helper utilities
# ---------------------------------------------------------------------------

def to_float(value):
    """Return float if value is numeric, else None."""
    if value is None:
        return None
    try:
        return float(str(value).strip())
    except (ValueError, TypeError):
        return None


def set_if_empty(ws, row, col, value):
    """Write value only if the cell is currently empty."""
    if ws.cell(row=row, column=col).value is None:
        ws.cell(row=row, column=col).value = value


def classify_t_test(est_type_str):
    """
    Return 'onesample', 'paired', 'independent', or None
    based on the free-text estimateType field.
    """
    s = str(est_type_str).lower()
    if "one-sample" in s or "one sample" in s:
        return "onesample"
    if "paired" in s:
        return "paired"
    if "independent" in s or "2-sample" in s or "two-sample" in s:
        return "independent"
    # plain 't-statistic' label — infer from df = n-1 pattern at call site
    return None


# ---------------------------------------------------------------------------
# Main calculation loop
# ---------------------------------------------------------------------------

def run_calculations(excel_path, sheet_name, overwrite=False):
    """
    Read the extraction sheet, compute missing effect sizes, and save.

    Parameters
    ----------
    excel_path : str   — path to the workbook
    sheet_name : str   — name of the data sheet
    overwrite  : bool  — if True, recalculate even if value already exists
                         (default False: skip rows that already have a value)
    """
    wb = openpyxl.load_workbook(excel_path)
    ws = wb[sheet_name]

    # Find columns dynamically by matching headers
    headers = [cell.value for cell in ws[1]]
    COL = {}
    for key, header_name in HEADER_MAP.items():
        if header_name in headers:
            COL[key] = headers.index(header_name) + 1
        else:
            raise ValueError(f"Header '{header_name}' not found in sheet '{sheet_name}'")

    calculated = []
    skipped    = []

    for row in range(2, ws.max_row + 1):

        study = ws.cell(row=row, column=COL["study"]).value
        if not study:
            continue

        exp       = ws.cell(row=row, column=COL["exp"]).value
        est_type  = str(ws.cell(row=row, column=COL["est_type"]).value  or "")
        data_type = str(ws.cell(row=row, column=COL["data_type"]).value or "")

        # Read numeric fields
        A  = to_float(ws.cell(row=row, column=COL["A"]).value)
        B  = to_float(ws.cell(row=row, column=COL["B"]).value)
        C  = to_float(ws.cell(row=row, column=COL["C"]).value)
        D  = to_float(ws.cell(row=row, column=COL["D"]).value)

        cm  = to_float(ws.cell(row=row, column=COL["ctrl_mean"]).value)
        csd = to_float(ws.cell(row=row, column=COL["ctrl_sd"]).value)
        cse = to_float(ws.cell(row=row, column=COL["ctrl_se"]).value)
        cn  = to_float(ws.cell(row=row, column=COL["ctrl_n"]).value)

        tm  = to_float(ws.cell(row=row, column=COL["treat_mean"]).value)
        tsd = to_float(ws.cell(row=row, column=COL["treat_sd"]).value)
        tse = to_float(ws.cell(row=row, column=COL["treat_se"]).value)
        tn  = to_float(ws.cell(row=row, column=COL["treat_n"]).value)

        t_val = to_float(ws.cell(row=row, column=COL["est_val"]).value)
        n_tot = to_float(ws.cell(row=row, column=COL["N"]).value)

        existing_hd = ws.cell(row=row, column=COL["hedges_d"]).value
        existing_or = ws.cell(row=row, column=COL["or"]).value

        done = []

        # ----------------------------------------------------------------
        # ODDS RATIO — requires positive A, B, C, D
        # ----------------------------------------------------------------
        if existing_or is None or overwrite:
            if all(v is not None and v > 0 for v in [A, B, C, D]):
                OR, var_or = odds_ratio_2x2(A, B, C, D)
                ws.cell(row=row, column=COL["or"]).value     = OR
                ws.cell(row=row, column=COL["or_var"]).value = var_or
                set_if_empty(ws, row, COL["or_how"],
                    f"OR=(A×D)/(B×C); Var(logOR)=1/A+1/B+1/C+1/D; "
                    f"A={A}, B={B}, C={C}, D={D}")
                done.append(f"OR={OR}, Var={var_or}")

        # ----------------------------------------------------------------
        # HEDGES' D — skip if already populated (unless overwrite=True)
        # ----------------------------------------------------------------
        if existing_hd is not None and not overwrite:
            if done:
                calculated.append((row, study, exp, done))
            continue

        # Case 1: from means + SD (SD may be recovered from SE * sqrt(n))
        csd_use = csd if csd is not None else (cse * math.sqrt(cn) if cse and cn else None)
        tsd_use = tsd if tsd is not None else (tse * math.sqrt(tn) if tse and tn else None)

        if all(v is not None for v in [cm, csd_use, cn, tm, tsd_use, tn]) \
                and cn > 0 and tn > 0:
            g, var_g = hedges_g_from_means(tm, tsd_use, int(tn),
                                           cm, csd_use, int(cn))
            ws.cell(row=row, column=COL["hedges_d"]).value     = g
            ws.cell(row=row, column=COL["hedges_d_var"]).value = var_g
            sd_note = "SD=SE×√n; " if (csd is None or tsd is None) else ""
            set_if_empty(ws, row, COL["hedges_d_how"],
                f"g=(treat−ctrl)/SP×J; {sd_note}"
                f"n_ctrl={int(cn)}, n_treat={int(tn)}")
            done.append(f"HedgesD={g} (from means)")

        # Case 2 / 3: from t-statistic
        elif t_val is not None and data_type == "Inferential statistics":

            t_class = classify_t_test(est_type)

            # Detect one-sample / paired by df = n-1 if label is ambiguous
            if t_class is None and n_tot is not None:
                # If est_extra mentions df = n-1, treat as one-sample
                extra_text = str(ws.cell(row=row, column=COL["est_extra"]).value or "")
                import re
                df_match = re.search(r"df\s*=\s*([0-9]+)", extra_text)
                if df_match:
                    df_reported = int(df_match.group(1))
                    if df_reported == int(n_tot) - 1:
                        t_class = "onesample"

            # Skip F-statistics and chi-squared (value in est_val is not a t)
            is_f_or_chi = any(kw in est_type.lower()
                              for kw in ("f-stat", "f stat", "chi", "glmm", "wald",
                                         "anova", "lm f"))
            if is_f_or_chi:
                skipped.append((row, study, exp,
                                 f"Non-t statistic: {est_type[:60]}"))
                if done:
                    calculated.append((row, study, exp, done))
                continue

            if t_class in ("onesample", "paired") and n_tot:
                n = int(n_tot)
                g, var_g = hedges_g_from_t_onesample(t_val, n)
                ws.cell(row=row, column=COL["hedges_d"]).value     = g
                ws.cell(row=row, column=COL["hedges_d_var"]).value = var_g
                set_if_empty(ws, row, COL["hedges_d_how"],
                    f"g=t/√n×J; {t_class} t={t_val}, n={n}")
                done.append(f"HedgesD={g} ({t_class} t)")

            elif t_class == "independent" and n_tot:
                n1 = n2 = int(n_tot) // 2   # equal split assumed if unknown
                g, var_g = hedges_g_from_t_indep(t_val, n1, n2)
                ws.cell(row=row, column=COL["hedges_d"]).value     = g
                ws.cell(row=row, column=COL["hedges_d_var"]).value = var_g
                set_if_empty(ws, row, COL["hedges_d_how"],
                    f"g=t×√(2/n)×J; independent t={t_val}, "
                    f"n1=n2={n1} (n1=n2=N/2 assumed)")
                done.append(f"HedgesD={g} (independent t)")

            else:
                skipped.append((row, study, exp,
                                 f"t-stat but could not classify: {est_type[:60]}"))

        if done:
            calculated.append((row, study, exp, done))
        elif existing_hd is None and existing_or is None:
            skipped.append((row, study, exp,
                             f"Insufficient data — dataType='{data_type}'"))

    # ----------------------------------------------------------------
    # CROSS-CONVERSION PASS
    # For each row: if one effect size is filled and the other is empty,
    # derive the missing one using the logistic approximation.
    # This is a FINAL pass so it never overwrites primary calculations.
    # ----------------------------------------------------------------
    converted = []

    for row in range(2, ws.max_row + 1):

        study = ws.cell(row=row, column=COL["study"]).value
        if not study:
            continue

        exp = ws.cell(row=row, column=COL["exp"]).value

        hd      = to_float(ws.cell(row=row, column=COL["hedges_d"]).value)
        hd_var  = to_float(ws.cell(row=row, column=COL["hedges_d_var"]).value)
        hd_how  = ws.cell(row=row, column=COL["hedges_d_how"]).value

        or_val  = to_float(ws.cell(row=row, column=COL["or"]).value)
        or_var  = to_float(ws.cell(row=row, column=COL["or_var"]).value)
        or_how  = ws.cell(row=row, column=COL["or_how"]).value

        # Best estimate of n for J correction (OR → g direction)
        n_tot = to_float(ws.cell(row=row, column=COL["N"]).value)
        A     = to_float(ws.cell(row=row, column=COL["A"]).value)
        B     = to_float(ws.cell(row=row, column=COL["B"]).value)
        if n_tot is not None and n_tot > 2:
            n_for_j = int(n_tot)
        elif A is not None and B is not None and (A + B) > 2:
            n_for_j = int(A + B)   # treatment group size as fallback
        else:
            n_for_j = None         # J ≈ 1

        # Case A: OR is filled, Hedges' D is blank → OR → g
        if (or_val is not None and or_var is not None
                and hd is None
                and or_val > 0):
            g, var_g, J = or_to_hedges_g(or_val, or_var, n=n_for_j)
            ws.cell(row=row, column=COL["hedges_d"]).value     = g
            ws.cell(row=row, column=COL["hedges_d_var"]).value = var_g
            j_note = f"J={round(J,4)}" if n_for_j else "J≈1 (n unknown)"
            ws.cell(row=row, column=COL["hedges_d_how"]).value = (
                f"[converted from OR] d=ln(OR)×√3/π×J; "
                f"OR={or_val}, Var(logOR)={or_var}, n={n_for_j}, {j_note}"
            )
            converted.append(
                (row, study, exp,
                 f"OR→HedgesD: g={g}, Var={var_g} ({j_note})")
            )

        # Case B: Hedges' D is filled, OR is blank → g → OR
        elif (hd is not None and hd_var is not None
              and or_val is None):
            OR, var_log_or = hedges_g_to_or(hd, hd_var)
            ws.cell(row=row, column=COL["or"]).value     = OR
            ws.cell(row=row, column=COL["or_var"]).value = var_log_or
            ws.cell(row=row, column=COL["or_how"]).value = (
                f"[converted from Hedges' D] logOR=g×π/√3; "
                f"g={hd}, Var(g)={hd_var}"
            )
            converted.append(
                (row, study, exp,
                 f"HedgesD→OR: OR={OR}, Var(logOR)={var_log_or}")
            )

    wb.save(excel_path)
    return calculated, skipped, converted


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":

    print("Running effect size calculations...")
    print(f"File: {EXCEL_PATH}\n")

    calculated, skipped, converted = run_calculations(
        excel_path=EXCEL_PATH,
        sheet_name=SHEET_NAME,
        overwrite=False,   # set True to recalculate all rows
    )

    print(f"{'='*60}")
    print(f"CALCULATED ({len(calculated)} rows)")
    print(f"{'='*60}")
    for row, study, exp, actions in calculated:
        print(f"  Row {row:2d}  {study[:35]}/{exp}")
        for a in actions:
            print(f"          {a}")

    print(f"\n{'='*60}")
    print(f"CROSS-CONVERTED ({len(converted)} rows)")
    print(f"{'='*60}")
    for row, study, exp, detail in converted:
        print(f"  Row {row:2d}  {study[:35]}/{exp}")
        print(f"          {detail}")

    print(f"\n{'='*60}")
    print(f"SKIPPED ({len(skipped)} rows)")
    print(f"{'='*60}")
    for row, study, exp, reason in skipped:
        print(f"  Row {row:2d}  {study[:35]}/{exp}")
        print(f"          Reason: {reason}")

    print("\nDone. File saved.")
